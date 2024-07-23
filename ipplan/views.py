from django.shortcuts import render
from django.shortcuts import redirect
from django.contrib import messages
from django.contrib.messages import constants
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.generic import CreateView
from django.views.generic import ListView
from django.views.generic import UpdateView
from django.views.generic import DeleteView
from django.views.generic import DetailView
from django.views.generic.edit import FormView
from django.http import JsonResponse
from ipplan.models import *
from ipplan.forms import *
from src.ipplan.func import *
import pandas
import ipaddress
import openpyxl

# Create your views here.


def validate_xss(list_string):
    regex = '<([A-Za-z_{}()/]+(\s|=)*)+>(.*<[A-Za-z/>]+)*'
    for string in list_string:
        result = re.search(regex, string)
        if result:
            return False
    return True

class RegionCreateView(CreateView):
    model = Region
    form_class = RegionForm
    template_name = "create_region.html"
    success_url = '/ipplan/create-region'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = self.request.user
        return super().form_valid(form)

class LocationCreateView(CreateView):
    model = Location
    form_class = LocationForm
    template_name = "create_location.html"
    success_url = '/ipplan/create-location'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = self.request.user
        return super().form_valid(form)

class SubnetCreateView(CreateView):
    model = Subnet
    form_class = SubnetForm
    template_name = "create_subnet.html"
    success_url = '/ipplan/list-subnet'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = self.request.user
        return super().form_valid(form)

class SubnetListView(ListView):
    model = Subnet
    context_object_name = 'subnets'
    template_name = "list_subnet.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
@login_required()
def request_ip_form(request):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if request.method == 'POST':
        form = RequestIpAddressForm(request.POST)
        if form.is_valid():
            subnet = form.cleaned_data.get('subnet')
            count = form.cleaned_data.get('count')
            status = form.cleaned_data.get('status')
            description = form.cleaned_data.get('description')
            used_ips = IpAddressModel.objects.filter(subnet__subnet=subnet).values_list('ip_address', flat=True)
            list_ips = get_available_ips(subnet=subnet, used_ips=used_ips, count=count)
            for ip in list_ips:
                model = IpAddressModel.objects.create(
                    ip_address = ip,
                    subnet = subnet,
                    status = status,
                    description = description,
                    user_created = request.user
                )
                model.save
            return render(request, template_name='request_ip_result.html', context={'list_ips': list_ips, 'subnet': subnet, 'status': status})
        else:
            return render(request, template_name='request_ip.html', context={'form': form})
    else:
        form = RequestIpAddressForm()
    return render(request, template_name='request_ip.html', context={'form': form})

@login_required()
def dashboard(request):
    return render(request, template_name='dashboard.html')

@login_required()
def list_ip(request, pk):
    list_ip = IpAddressModel.objects.filter(subnet__id=pk)
    return render(request, template_name='list_ip.html', context={'list_ip': list_ip})


class IpAddressModelDeleteView(DeleteView):
    model = IpAddressModel
    template_name = 'list_ip.html'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_object(self):
        return IpAddressModel.objects.get(pk=self.kwargs["id"])

    def get_success_url(self):
        pk = self.kwargs["pk"]
        pk1 = self.kwargs['id']
        return reverse("list-ip", kwargs={"pk": pk})


class IpAddressModelUpdateView(UpdateView):
    model = IpAddressModel
    form_class = IpAddressModelUpdatelForm
    template_name = "update_ip.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_object(self):
        return IpAddressModel.objects.get(pk=self.kwargs["id"])

    def get_success_url(self):
        pk = self.kwargs["pk"]
        pk1 = self.kwargs['id']
        return reverse("list-ip", kwargs={"pk": pk})


class SubnetUpdateView(UpdateView):
    model = Subnet
    form_class = SubnetForm
    template_name = "update_subnet.html"
    success_url = '/ipplan/list-subnet'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class IpAddressModelDetailView(DetailView):
    model = IpAddressModel
    form_class = IpAddressModelForm
    template_name = "detail_ip.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_object(self):
        return IpAddressModel.objects.get(pk=self.kwargs["id"])


class SubnetDeleteView(DeleteView):
    model = Subnet
    template_name = "list-subnet.html"
    success_url = '/ipplan/list-subnet'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)


def create_multiple_subnet(request):
    try:
        if request.method == 'POST' and request.FILES.get('upload-file'):
            uploaded_file = request.FILES['upload-file']
            wb = openpyxl.load_workbook(uploaded_file)
            worksheet = wb["IPPlan"]
            for item in worksheet.iter_rows(min_row=2):
                item = ["" if i.value is None else i.value for i in item]
                region = item[0]
                location = item[1]
                subnet = item[2]
                name = item[3]
                description = item[4]
                validate_xss_result = validate_xss(list_string=item)
                if validate_xss_result:
                    if region != "" or location != "" or subnet != "":
                        obj_count_01 = Region.objects.filter(region=region).count()
                        obj_count_02 = Location.objects.filter(location=location).count()
                        obj_count_03 = Subnet.objects.filter(subnet=subnet).count()
                        if obj_count_01 == 0:
                            model = Region(
                                region=region,
                                description=region,
                                user_created=request.user
                            )
                            model.save()
                        obj_01 = Region.objects.get(region=region)
                        if obj_count_02 == 0:
                            model = Location(
                                location=location,
                                region=obj_01,
                                description=region,
                                user_created=request.user
                            )
                            model.save()
                        if obj_count_03 == 0:
                            obj_02 = Location.objects.get(location=location)
                            model = Subnet(
                                subnet=subnet,
                                location=obj_02,
                                name=name,
                                description=description,
                                user_created=request.user
                            )
                            model.save()
                        else:
                            model = Subnet.objects.get(subnet=subnet)
                            model.location = Location.objects.get(location=location)
                            model.name = name
                            model.description = description
                            model.user_created = str(request.user)
                            model.save()
                    messages.add_message(request, constants.SUCCESS, 'Import subnet success')
                else:
                    raise Exception('The input string contains unusual characters')
    except Exception as error:
        messages.add_message(request, constants.ERROR, f'An error occurred: {error}')
    return render(request, template_name='create_multiple_subnet.html')


def request_multiple_ip(request):
    try:
        if request.method == 'POST' and request.FILES.get('upload-file'):
            uploaded_file = request.FILES['upload-file']
            wb = openpyxl.load_workbook(uploaded_file)
            worksheet = wb["DATA"]
            list_subnet = Subnet.objects.all().values_list('subnet', flat=True)
            for item in worksheet.iter_rows(min_row=2):
                item = ["" if i.value is None else i.value for i in item]
                description = item[0]
                ip = item[1]
                status = item[2]
                subnet = [subnet for subnet in list_subnet if is_ip(ip) is True and is_subnet(subnet) is True and ipaddress.IPv4Address(ip) in ipaddress.ip_network(subnet)]
                get_ip_status = IpStatus.objects.filter(status=status).count()
                if subnet and get_ip_status > 0:
                    subnet_obj = Subnet.objects.get(subnet=subnet[-1])
                    status_obj = IpStatus.objects.get(status=status)
                    IpAddressModel.objects.update_or_create(
                        ip_address=ip,
                        defaults={
                            'status': status_obj,
                            'subnet': subnet_obj,
                            'description': description
                        }
                    )
            messages.add_message(request, constants.SUCCESS, 'Import IP success')
    except Exception as error:
        messages.add_message(request, constants.ERROR, f'An error occurred: {error}')
    return render(request, template_name='request_multiple_ip.html')

def list_subnet_tree(request):
    context = {}
    if request.method == 'GET':
        regions = Region.objects.all()
        locations = Location.objects.all()
        subnets = Subnet.objects.all()
        ips = IpAddressModel.objects.all()
        context = {'regions': regions, 'locations': locations, 'subnets': subnets, 'ips': ips}
    return render(request, 'list_subnet_tree.html', context=context)