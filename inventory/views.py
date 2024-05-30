from django.db.models import ProtectedError
from django.shortcuts import render
from django.shortcuts import redirect
from django.contrib import messages
from django.contrib.messages import constants
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.generic import CreateView
from django.views.generic import ListView
from django.views.generic import UpdateView
from django.views.generic import DeleteView
from django.views.generic import DetailView
from django.conf import settings
from inventory.models import *
from inventory.forms import *
import pandas
import datetime
import openpyxl
# Create your views here.

MESSAGE_TAGS = {
    messages.ERROR: 'danger'
}


def validate_xss(list_string):
    regex = '<([A-Za-z_{}()/]+(\s|=)*)+>(.*<[A-Za-z/>]+)*'
    for string in list_string:
        result = re.search(regex, string)
        if result:
            return False
    return True

class DeviceCreateView(CreateView):
    model = Device
    form_class = DeviceForm
    template_name = "create_device.html"
    success_url = '/inventory/list-device/device'
    
    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Create success')
        return super().form_valid(form)

class DeviceProvinceCreateView(CreateView):
    model = DeviceProvince
    form_class = DeviceProvinceForm
    template_name = "create_device_province.html"
    success_url = '/inventory/list-device-province'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Create success')
        return super().form_valid(form)

class DeviceProvinceUpdateView(UpdateView):
    model = DeviceProvince
    form_class = DeviceProvinceForm
    template_name = "update_device_province.html"
    success_url = '/inventory/list-device-province'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Update success')
        return super().form_valid(form)

class DeviceProvinceDeleteView(DeleteView):
    model = DeviceProvince
    template_name = 'list_device_province.html'
    success_url = '/inventory/list-device-province'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(self.request, constants.SUCCESS, 'Delete success')
        except ProtectedError:
            messages.add_message(self.request, constants.ERROR, 'This object has been protected')
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)

class DeviceProvinceListView(ListView):
    model = DeviceProvince
    context_object_name = 'objects'
    template_name = "list_device_province.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class DeviceTypeCreateView(CreateView):
    model = DeviceType
    form_class = DeviceTypeForm
    template_name = "create_device_type.html"
    success_url = '/inventory/list-device-type'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Create success')
        return super().form_valid(form)

class DeviceTypeListView(ListView):
    model = DeviceType
    context_object_name = 'objects'
    template_name = "list_device_type.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class DeviceTypeUpdateView(UpdateView):
    model = DeviceType
    form_class = DeviceTypeForm
    template_name = "update_device_type.html"
    success_url = '/inventory/list-device-type'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Update success')
        return super().form_valid(form)

class DeviceTypeDeleteView(DeleteView):
    model = DeviceType
    template_name = 'list_device_type.html'
    success_url = '/inventory/list-device-type'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(self.request, constants.SUCCESS, 'Delete success')
        except ProtectedError:
            messages.add_message(self.request, constants.ERROR, 'This object has been protected')
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)

class DeviceCategoryCreateView(CreateView):
    model = DeviceCategory
    form_class = DeviceCategoryForm
    template_name = "create_device_category.html"
    success_url = 'list-device-category'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Create success')
        return super().form_valid(form)

class DeviceCategoryListView(ListView):
    model = DeviceCategory
    context_object_name = 'objects'
    template_name = "list_device_category.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class DeviceCategoryUpdateView(UpdateView):
    model = DeviceCategory
    form_class = DeviceCategoryForm
    template_name = "update_device_category.html"
    success_url = '/inventory/list-device-category'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Update success')
        return super().form_valid(form)

class DeviceCategoryDeleteView(DeleteView):
    model = DeviceCategory
    template_name = 'list_device_category.html'
    success_url = '/inventory/list-device-category'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(self.request, constants.SUCCESS, 'Delete success')
        except ProtectedError:
            messages.add_message(self.request, constants.ERROR, 'This object has been protected')
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)

class DeviceVendorCreateView(CreateView):
    model = DeviceVendor
    form_class = DeviceVendorForm
    template_name = "create_device_vendor.html"
    success_url = '/inventory/list-device-vendor'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Create success')
        return super().form_valid(form)

class DeviceVendorListView(ListView):
    model = DeviceVendor
    context_object_name = 'objects'
    template_name = "list_device_vendor.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class DeviceVendorUpdateView(UpdateView):
    model = DeviceVendor
    form_class = DeviceVendorForm
    template_name = "update_device_vendor.html"
    success_url = '/inventory/list-device-vendor'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Update success')
        return super().form_valid(form)

class DeviceVendorDeleteView(DeleteView):
    model = DeviceVendor
    template_name = 'list_device_vendor.html'
    success_url = '/inventory/list-device-vendor'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(self.request, constants.SUCCESS, 'Delete success')
        except ProtectedError:
            messages.add_message(self.request, constants.ERROR, 'This object has been protected')
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)

class DeviceListView(ListView):
    model = Device
    context_object_name = 'devices'
    template_name = "list_device.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class DeviceUpdateView(UpdateView):
    model = Device
    form_class = DeviceForm
    template_name = "update_device.html"
    success_url = '/inventory/list-device/device'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Update success')
        return super().form_valid(form)

class DeviceDeleteView(DeleteView):
    model = Device
    template_name = 'list_device.html'
    success_url = '/inventory/list-device/device'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(self.request, constants.SUCCESS, 'Delete success')
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)

class DeviceDetailView(DetailView):
    model = Device
    template_name = "detail_device.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
@login_required()
def create_multiple_device(request):
    try:
        if request.method == 'POST' and request.FILES.get('upload-file'):
            uploaded_file = request.FILES['upload-file']
            wb = openpyxl.load_workbook(uploaded_file)
            worksheet_01 = wb["Device"]
            worksheet_02 = wb["Device_management"]
            worksheet_03 = wb["Device_rack_layout"]
            for item in worksheet_01.iter_rows(min_row=2, values_only=True):
                item = ["" if i is None else i for i in item]
                obj_count = Device.objects.filter(device_ip=item[1]).count()
                validate_xss_result = validate_xss(list_string=item)
                if validate_xss_result:
                    if obj_count == 0:
                        obj_count_01 = DeviceBranch.objects.filter(device_branch=item[2]).count()
                        obj_count_02 = DeviceProvince.objects.filter(device_province=item[3]).count()
                        obj_count_03 = DeviceType.objects.filter(device_type=item[4]).count()
                        obj_count_04 = DeviceCategory.objects.filter(device_category=item[5]).count()
                        obj_count_05 = DeviceVendor.objects.filter(device_vendor=item[6]).count()
                        obj_count_06 = DeviceOS.objects.filter(device_os=item[7]).count()
                        obj_count_07 = DeviceTag.objects.filter(device_tag=item[8]).count()
                        if obj_count_01 == 0:
                            model_01 = DeviceBranch(device_branch=item[2])
                            model_01.save()
                        if obj_count_02 == 0:
                            model_02 = DeviceProvince(device_province=item[3])
                            model_02.save()
                        if obj_count_03 == 0:
                            model_03 = DeviceType(device_type=item[4])
                            model_03.save()
                        if obj_count_04 == 0:
                            model_04 = DeviceCategory(device_category=item[5])
                            model_04.save()
                        if obj_count_05 == 0:
                            model_05 = DeviceVendor(device_vendor=item[6])
                            model_05.save()
                        if obj_count_06 == 0:
                            model_06 = DeviceOS(device_os=item[7])
                            model_06.save()
                        if obj_count_07 == 0:
                            model_07 = DeviceTag(device_tag=item[8])
                            model_07.save()
                        model = Device(
                            device_name=item[0],
                            device_ip=item[1],
                            device_branch=DeviceBranch.objects.get(device_branch=item[2]),
                            device_province=DeviceProvince.objects.get(device_province=item[3]),
                            device_type=DeviceType.objects.get(device_type=item[4]),
                            device_category=DeviceCategory.objects.get(device_category=item[5]),
                            device_vendor=DeviceVendor.objects.get(device_vendor=item[6]),
                            device_os=DeviceOS.objects.get(device_os=item[7]),
                            device_tag=DeviceTag.objects.get(device_tag=item[8]),
                            device_stack=item[9],
                            device_description=item[10],
                            user_created=request.user
                        )
                        model.save()
                    elif obj_count == 1:
                        obj_count_01 = DeviceBranch.objects.filter(device_branch=item[2]).count()
                        obj_count_02 = DeviceProvince.objects.filter(device_province=item[3]).count()
                        obj_count_03 = DeviceType.objects.filter(device_type=item[4]).count()
                        obj_count_04 = DeviceCategory.objects.filter(device_category=item[5]).count()
                        obj_count_05 = DeviceVendor.objects.filter(device_vendor=item[6]).count()
                        obj_count_06 = DeviceOS.objects.filter(device_os=item[7]).count()
                        obj_count_07 = DeviceTag.objects.filter(device_tag=item[8]).count()
                        if obj_count_01 == 0:
                            model_01 = DeviceBranch(device_branch=item[2])
                            model_01.save()
                        if obj_count_02 == 0:
                            model_02 = DeviceProvince(device_province=item[3])
                            model_02.save()
                        if obj_count_03 == 0:
                            model_03 = DeviceType(device_type=item[4])
                            model_03.save()
                        if obj_count_04 == 0:
                            model_04 = DeviceCategory(device_category=item[5])
                            model_04.save()
                        if obj_count_05 == 0:
                            model_05 = DeviceVendor(device_vendor=item[6])
                            model_05.save()
                        if obj_count_06 == 0:
                            model_06 = DeviceOS(device_os=item[7])
                            model_06.save()
                        if obj_count_07 == 0:
                            model_07 = DeviceTag(device_tag=item[8])
                            model_07.save()
                        model = Device.objects.get(device_ip=item[1])
                        model.device_name = item[0]
                        model.device_branch = DeviceBranch.objects.get(device_branch=item[2])
                        model.device_province = DeviceProvince.objects.get(device_province=item[3])
                        model.device_type = DeviceType.objects.get(device_type=item[4])
                        model.device_category = DeviceCategory.objects.get(device_category=item[5])
                        model.device_vendor = DeviceVendor.objects.get(device_vendor=item[6])
                        model.device_os = DeviceOS.objects.get(device_os=item[7])
                        model.device_tag = DeviceTag.objects.get(device_tag=item[8])
                        model.device_stack = item[9]
                        model.device_description = item[10]
                        model.user_created = str(request.user)
                        model.save()
                else:
                    raise Exception('The input string contains unusual characters')
            for item in worksheet_02.iter_rows(min_row=2, values_only=True):
                validate_xss_result = validate_xss(list_string=item)
                if validate_xss_result:
                    if item[0] is not None and item[1] is not None:
                        obj_count = Device.objects.filter(device_ip=item[1]).count()
                        if obj_count == 1:
                            device_tag = Device.objects.get(device_ip=item[1]).device_tag
                            obj_count_01 = DeviceManagement.objects.filter(device_ip=Device.objects.get(device_ip=item[1])).count()
                            obj_count_02 = DeviceManagement.objects.filter(device_serial_number=item[2]).count()
                            if obj_count_01 == 0 or (obj_count_01 == 1 and device_tag == 'duplicated'):
                                if obj_count_02 == 0:
                                    model = DeviceManagement(
                                        device_ip=Device.objects.get(device_ip=item[1]),
                                        device_serial_number=item[2],
                                        start_ma_date=item[3],
                                        end_ma_date=item[4],
                                        start_license_date=item[5],
                                        end_license_date=item[6],
                                        end_sw_support_date=item[7],
                                        end_hw_support_date=item[8],
                                        start_used_date=item[9],
                                        user_created=request.user
                                    )
                                    model.save()
                else:
                    raise Exception('The input string contains unusual characters')
            for item in worksheet_03.iter_rows(min_row=2, values_only=True):
                validate_xss_result = validate_xss(list_string=item)
                if validate_xss_result:
                    if item[0] is not None and item[1] is not None:
                        obj_count = Device.objects.filter(device_ip=item[1]).count()
                        if obj_count != 0:
                            device_tag = Device.objects.get(device_ip=item[1]).device_tag
                            obj_count_01 = DeviceRackLayout.objects.filter(device_ip=Device.objects.get(device_ip=item[1])).count()
                            if obj_count_01 == 0:
                                model = DeviceRackLayout(
                                    device_ip=Device.objects.get(device_ip=item[1]),
                                    device_rack_name=item[2],
                                    device_rack_unit=item[3],
                                    user_created=request.user
                                )
                                model.save()
                            elif obj_count_01 == 1 and device_tag == 'duplicated':
                                obj = DeviceRackLayout.objects.get(device_ip=Device.objects.get(device_ip=item[1]))
                                rack_name = obj.device_rack_name
                                rack_unit = obj.device_rack_unit
                                if rack_name != item[2] or rack_unit != item[3]:
                                    model = DeviceRackLayout(
                                        device_ip=Device.objects.get(device_ip=item[1]),
                                        device_rack_name=item[2],
                                        device_rack_unit=item[3],
                                        user_created=request.user
                                    )
                                    model.save()
                else:
                    raise Exception('The input string contains unusual characters')
            messages.add_message(request, constants.SUCCESS, 'Import devices success')
    except Exception as error:
        messages.add_message(request, constants.ERROR, f'An error occurred: {error}')
    return render(request, 'create_multiple_device.html')

@login_required()
def device_dashboard(request):
    return render(request, template_name='device_dashboard.html')

class DeviceManagementListView(ListView):
    model = DeviceManagement
    form_class = DeviceManagementForm
    context_object_name = 'devices'
    template_name = "list_device_management.html"
    
    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class DeviceManagementCreateView(CreateView):
    model = DeviceManagement
    form_class = DeviceManagementForm
    template_name = "create_device_management.html"
    success_url = '/inventory/list-device/device-management'
    
    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Create success')
        return super().form_valid(form)
    
class DeviceManagementDeleteView(DeleteView):
    model = DeviceManagement
    template_name = 'list_device.html'
    success_url = '/inventory/list-device/device-management'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(self.request, constants.SUCCESS, 'Delete success')
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)

class DeviceInterfaceListView(ListView):
    model = DeviceInterface
    form_class = DeviceInterfaceForm
    context_object_name = 'devices'
    template_name = "list_device_interface.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class DeviceRackLayoutListView(ListView):
    model = DeviceRackLayout
    form_class = DeviceRackLayoutForm
    context_object_name = 'devices'
    template_name = "list_device_rack_layout.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class DeviceRackLayoutCreateView(CreateView):
    model = DeviceRackLayout
    form_class = DeviceRackLayoutForm
    template_name = "create_device_rack_layout.html"
    success_url = '/inventory/list-device/device-rack-layout'
    
    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Create success')
        return super().form_valid(form)

class DeviceConfigurationListView(ListView):
    model = DeviceConfiguration
    form_class = DeviceConfigurationForm
    context_object_name = 'devices'
    template_name = "list_device_configuration.html"
    
    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class DeviceInterfaceDetailView(DetailView):
    model = DeviceInterface
    template_name = "detail_device_interface.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class DeviceManagementUpdateView(UpdateView):
    model = DeviceManagement
    form_class = DeviceManagementForm
    template_name = 'update_device_management.html'
    success_url = '/inventory/list-device/device-management'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Update success')
        return super().form_valid(form)
    
class DeviceManagementDetailView(DetailView):
    model = DeviceManagement
    template_name = 'detail_device_management.html'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
class DeviceRackLayoutUpdateView(UpdateView):
    model = DeviceRackLayout
    form_class = DeviceRackLayoutForm
    template_name = 'update_device_rack_layout.html'
    success_url = '/inventory/list-device/device-rack-layout'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Update success')
        return super().form_valid(form)
    
class DeviceRackLayoutDetailView(DetailView):
    model = DeviceRackLayout
    template_name = 'detail_device_rack_layout.html'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
class DeviceRackLayoutDeleteView(DeleteView):
    model = DeviceRackLayout
    template_name = 'list_device_rack_layout.html'
    success_url = '/inventory/list-device/device-rack-layout'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(self.request, constants.SUCCESS, 'Delete success')
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)

class DeviceOSCreateView(CreateView):
    model = DeviceOS
    form_class = DeviceOSForm
    template_name = "create_device_os.html"
    success_url = '/inventory/list-device-os'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Create success')
        return super().form_valid(form)

class DeviceOSListView(ListView):
    model = DeviceOS
    context_object_name = 'objects'
    template_name = "list_device_os.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class DeviceOSUpdateView(UpdateView):
    model = DeviceOS
    form_class = DeviceOSForm
    template_name = "update_device_os.html"
    success_url = '/inventory/list-device-os'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Update success')
        return super().form_valid(form)

class DeviceOSDeleteView(DeleteView):
    model = DeviceOS
    template_name = 'list_device_os.html'
    success_url = '/inventory/list-device-os'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(self.request, constants.SUCCESS, 'Delete success')
        except ProtectedError:
            messages.add_message(self.request, constants.ERROR, 'This object has been protected')
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)
    
def device_export(request):
    form = DeviceExportForm
    data = {'form': form}
    try:
        if request.method == 'POST':
            form = DeviceExportForm(request.POST)
            if form.is_valid():
                select_id = form.data['database_table']
                if select_id == '1':
                    datalist = list()
                    queryset = Device.objects.all()
                    for item in queryset:
                        datalist.append({
                            'device_name': item.device_name,
                            'device_ip': item.device_ip,
                            'device_branch': item.device_branch,
                            'device_province': item.device_province,
                            'device_type': item.device_type,
                            'device_category': item.device_category,
                            'device_vendor': item.device_vendor,
                            'device_os': item.device_os,
                            'device_firmware': item.device_firmware,
                            'device_stack': item.device_stack,
                            'device_tag': item.device_tag,
                            'device_description': item.device_description,
                            'device_creation_time': str(item.device_creation_time),
                            'user_created': item.user_created
                        })
                    df = pandas.DataFrame(datalist)
                    df.to_csv(settings.MEDIA_ROOT + '/inventory/device.csv', encoding='utf-8-sig')
                    download_url = '/media/inventory/device.csv'
                    messages.add_message(request, constants.SUCCESS, download_url)
                elif select_id == '2':
                    datalist = list()
                    queryset = DeviceManagement.objects.all()
                    for item in queryset:
                        datalist.append({
                            'device_branch': item.device_ip.device_branch,
                            'device_province': item.device_ip.device_province,
                            'device_name': item.device_ip,
                            'device_ip': item.device_ip,
                            'device_serial_number': item.device_serial_number,
                            'start_ma_date': str(item.start_ma_date),
                            'end_ma_date': str(item.end_ma_date),
                            'start_license_date': str(item.start_license_date),
                            'end_license_date': str(item.end_license_date),
                            'end_sw_support_date': str(item.end_sw_support_date),
                            'end_hw_support_date': str(item.end_hw_support_date),
                            'start_used_date': str(item.start_used_date),
                            'user_created': item.user_created
                        })
                    df = pandas.DataFrame(datalist)
                    df.to_csv(settings.MEDIA_ROOT + '/inventory/device-management.csv', encoding='utf-8-sig')
                    download_url = '/media/inventory/device-management.csv'
                    messages.add_message(request, constants.SUCCESS, download_url)
                elif select_id == '3':
                    datalist = list()
                    queryset = DeviceRackLayout.objects.all()
                    for item in queryset:
                        datalist.append({
                            'device_branch': item.device_ip.device_branch,
                            'device_province': item.device_ip.device_province,
                            'device_name': item.device_ip.device_name,
                            'device_ip': item.device_ip,
                            'device_rack_name': item.device_rack_name,
                            'device_rack_unit': item.device_rack_unit,
                            'user_created': item.user_created
                        })
                    df = pandas.DataFrame(datalist)
                    df.to_csv(settings.MEDIA_ROOT + '/inventory/device-rack-layout.csv', encoding='utf-8-sig')
                    download_url = '/media/inventory/device-rack-layout.csv'
                    messages.add_message(request, constants.SUCCESS, download_url)
                elif select_id == '4':
                    datalist = list()
                    queryset = DeviceConfiguration.objects.all()
                    for item in queryset:
                        datalist.append({
                            'device_branch': item.device_ip.device_branch,
                            'device_province': item.device_ip.device_province,
                            'device_name': item.device_ip.device_name,
                            'device_ip': item.device_ip,
                            'device_config_standardized': item.device_config_standardized,
                            'device_monitored': item.device_monitored,
                            'device_backup_config': item.device_backup_config,
                            'user_created': item.user_created
                        })
                    df = pandas.DataFrame(datalist)
                    df.to_csv(settings.MEDIA_ROOT + '/inventory/device-configuration.csv', encoding='utf-8-sig')
                    download_url = '/media/inventory/device-configuration.csv'
                    messages.add_message(request, constants.SUCCESS, download_url)
                elif select_id == '5':
                    list_device_firmware = Device.objects.all().values_list('device_name', 'device_ip', 'device_type__device_type', 'device_firmware')
                    list_firmware = DeviceFirmware.objects.all().values_list('device_type__device_type', 'firmware')
                    datalist = list()
                    for item in list_device_firmware:
                        device_name = item[0]
                        device_ip = item[1]
                        device_type = item[2]
                        device_firmware = item[3]
                        if device_firmware is not None:
                            checklist = [i for i in list_firmware if i == (device_type, device_firmware)]
                            if not checklist:
                                datalist.append({
                                    'device_name': device_name,
                                    'device_ip': device_ip,
                                    'device_type': device_type,
                                    'device_firmware': device_firmware
                                })
                    df = pandas.DataFrame(datalist)
                    df.to_csv(settings.MEDIA_ROOT + '/inventory/device-incorrect-firmware.csv', encoding='utf-8-sig')
                    download_url = '/media/inventory/device-incorrect-firmware.csv'
                    messages.add_message(request, constants.SUCCESS, download_url)
                elif select_id == '6':
                    datepoint01 = datetime.date.today()
                    datepoint02 = datetime.date.today() + datetime.timedelta(days=180)
                    list_device_serial = DeviceManagement.objects.all().values_list('device_serial_number', flat=True)
                    dataset01 = list()
                    dataset02 = list()
                    dataset03 = list()
                    dataset04 = list()
                    for device_serial in list_device_serial:                     
                        datalist01 = DeviceManagement.objects.filter(device_serial_number=device_serial, end_ma_date__gte=datepoint01, end_ma_date__lte=datepoint02)
                        datalist02 = DeviceManagement.objects.filter(device_serial_number=device_serial, end_license_date__gte=datepoint01, end_license_date__lte=datepoint02)
                        datalist03 = DeviceManagement.objects.filter(device_serial_number=device_serial, end_sw_support_date__gte=datepoint01, end_sw_support_date__lte=datepoint02)
                        datalist04 = DeviceManagement.objects.filter(device_serial_number=device_serial, end_hw_support_date__gte=datepoint01, end_hw_support_date__lte=datepoint02)
                        if datalist01:
                            for item in datalist01:
                                dataset01.append({
                                    'device_branch': item.device_ip.device_branch,
                                    'device_province': item.device_ip.device_province,
                                    'device_name': item.device_ip.device_name,
                                    'device_ip': item.device_ip,
                                    'device_serial': item.device_serial_number,
                                    'end_ma': item.end_ma_date
                                })
                        if datalist02:
                            for item in datalist02:
                                dataset02.append({
                                    'device_branch': item.device_ip.device_branch,
                                    'device_province': item.device_ip.device_province,
                                    'device_name': item.device_ip.device_name,
                                    'device_ip': item.device_ip,
                                    'device_serial': item.device_serial_number,
                                    'end_license': item.end_license_date
                                })
                        if datalist03:
                            for item in datalist03:
                                dataset03.append({
                                    'device_branch': item.device_ip.device_branch,
                                    'device_province': item.device_ip.device_province,
                                    'device_name': item.device_ip.device_name,
                                    'device_ip': item.device_ip,
                                    'device_serial': item.device_serial_number,
                                    'end_hw_support': item.end_sw_support_date
                                })
                        if datalist04:
                            for item in datalist04:
                                dataset04.append({
                                    'device_branch': item.device_ip.device_branch,
                                    'device_province': item.device_ip.device_province,
                                    'device_name': item.device_ip.device_name,
                                    'device_ip': item.device_ip,
                                    'device_serial': item.device_serial_number,
                                    'end_sw_support': item.end_hw_support_date
                                })
                    df01 = pandas.DataFrame(dataset01)
                    df02 = pandas.DataFrame(dataset02)
                    df03 = pandas.DataFrame(dataset03)
                    df04 = pandas.DataFrame(dataset04)
                    with pandas.ExcelWriter(settings.MEDIA_ROOT + '/inventory/device-expired-license.xlsx') as w:
                        df01.to_excel(w, sheet_name="End MA")
                        df02.to_excel(w, sheet_name="End License")
                        df03.to_excel(w, sheet_name="End SW support")
                        df04.to_excel(w, sheet_name="End HW support")
                    download_url = '/media/inventory/device-expired-license.xlsx'
                    messages.add_message(request, constants.SUCCESS, download_url)
                else:
                    messages.add_message(request, constants.ERROR, 'Database selected is not validate')
    except Exception as error:
        messages.add_message(request, constants.ERROR, error)
    return render(request, template_name='device_export.html', context=data)

class DeviceFirmwareCreateView(CreateView):
    model = DeviceFirmware
    form_class = DeviceFirmwareForm
    template_name = "create_device_firmware.html"
    success_url = '/inventory/list-device-firmware'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Create success')
        return super().form_valid(form)

class DeviceFirmwareListView(ListView):
    model = DeviceFirmware
    context_object_name = 'objects'
    template_name = "list_device_firmware.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

class DeviceFirmwareUpdateView(UpdateView):
    model = DeviceFirmware
    form_class = DeviceFirmwareForm
    template_name = "update_device_firmware.html"
    success_url = '/inventory/list-device-firmware'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Update success')
        return super().form_valid(form)

class DeviceFirmwareDeleteView(DeleteView):
    model = DeviceFirmware
    template_name = 'list_device_firmware.html'
    success_url = '/inventory/list-device-firmware'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(self.request, constants.SUCCESS, 'Delete success')
        except ProtectedError:
            messages.add_message(self.request, constants.ERROR, 'This object has been protected')
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)
    
class DeviceBranchCreateView(CreateView):
    model = DeviceBranch
    form_class = DeviceBranchForm
    template_name = "create_device_branch.html"
    success_url = '/inventory/list-device-branch'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Create success')
        return super().form_valid(form)

class DeviceBranchUpdateView(UpdateView):
    model = DeviceBranch
    form_class = DeviceBranchForm
    template_name = "update_device_branch.html"
    success_url = '/inventory/list-device-branch'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Update success')
        return super().form_valid(form)

class DeviceBranchDeleteView(DeleteView):
    model = DeviceBranch
    template_name = 'list_device_branch.html'
    success_url = '/inventory/list-device-branch'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(self.request, constants.SUCCESS, 'Delete success')
        except ProtectedError:
            messages.add_message(self.request, constants.ERROR, 'This object has been protected')
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)

class DeviceBranchListView(ListView):
    model = DeviceBranch
    context_object_name = 'objects'
    template_name = "list_device_branch.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    

class DeviceTagCreateView(CreateView):
    model = DeviceTag
    form_class = DeviceTagForm
    template_name = "create_device_tag.html"
    success_url = '/inventory/list-device-tag'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Create success')
        return super().form_valid(form)

class DeviceTagUpdateView(UpdateView):
    model = DeviceTag
    form_class = DeviceTagForm
    template_name = "update_device_tag.html"
    success_url = '/inventory/list-device-tag'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        form.instance.user_created = str(self.request.user)
        messages.add_message(self.request, constants.SUCCESS, 'Update success')
        return super().form_valid(form)

class DeviceTagDeleteView(DeleteView):
    model = DeviceTag
    template_name = 'list_device_tag.html'
    success_url = '/inventory/list-device-tag'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            super().post(request, *args, **kwargs)
            messages.add_message(self.request, constants.SUCCESS, 'Delete success')
        except ProtectedError:
            messages.add_message(self.request, constants.ERROR, 'This object has been protected')
        except Exception as error:
            messages.add_message(self.request, constants.ERROR, error)
        return redirect(self.success_url)

class DeviceTagListView(ListView):
    model = DeviceTag
    context_object_name = 'objects'
    template_name = "list_device_tag.html"

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)